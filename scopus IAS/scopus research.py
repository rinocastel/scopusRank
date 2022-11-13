import string
import requests
from openpyxl import load_workbook, Workbook

import json
import sys
import os
import xlsxwriter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.service import Service
#options = webdriver.ChromeOptions()
#driver = webdriver.Chrome(executable_path="C:/Users/Rino/Documents/Lavoro/Unipd/Scopus/scopus IAS/chromedriver.exe", options=options)
f=open(os.getcwd()+"\\ISBN_IAS.txt",'r')
lines=f.readlines()
f.close()

HEADERS=["Title","Authors","Publication Name","Count Citation"]
dest_wb = Workbook()




query_scopus = {} 
query_scopus['count']=24
for line in lines:
    dest_ws = dest_wb.create_sheet(title="IAS-"+line.split(',')[1])
    dest_ws.append(HEADERS)
    #facets = f"https://api.elsevier.com/content/search/scopus?query=isbn(1586030787)&sort=citedby-count"
    facets = f"https://api.elsevier.com/content/search/scopus?query=isbn("+str(line.split(',')[0])+")&sort=citedby-count"
    resp2 = requests.get(facets,params=query_scopus,
                    headers={'Accept':'application/json','X-ELS-APIKey': '34238e4e4a2ba14959cc54c5d1c501d7'})
    cont=0
    rows=[]
    print(resp2.json())
    for r in resp2.json()['search-results']['entry']:
        
        print("-----------"+str(cont))
        cont+=1
        try:
            rows.append(tuple((r['dc:title'],r['dc:creator'],r['prism:publicationName'],r['citedby-count'])))
        except:
            rows.append(tuple((r['dc:title'],'/',r['prism:publicationName'],r['citedby-count'])))
    rows.sort(key=lambda tup: int(tup[3]))
    rows.reverse()
    for row in rows:
        r=[]
        r.append(row[0])
        r.append(row[1])
        r.append(row[2])
        r.append(row[3])
        dest_ws.append(r)
dest_wb.save(filename="IAS-RANK.xlsx")    
dest_wb.close()